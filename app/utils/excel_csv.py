from typing import List, Dict, IO
import csv
from io import BytesIO, StringIO
try:
    import openpyxl
except Exception:
    openpyxl = None

def parse_csv(content: bytes) -> List[Dict]:
    text = content.decode('utf-8', errors='ignore')
    f = StringIO(text)
    reader = csv.DictReader(f)
    return [dict(row) for row in reader]

def parse_xlsx(content: bytes) -> List[Dict]:
    if openpyxl is None:
        raise RuntimeError('openpyxl 未安装')
    wb = openpyxl.load_workbook(filename=BytesIO(content), read_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2):
        obj = {}
        for i, c in enumerate(row):
            key = headers[i] if i < len(headers) else f'col{i}'
            obj[key] = c.value if c.value is not None else ''
        rows.append(obj)
    return rows

